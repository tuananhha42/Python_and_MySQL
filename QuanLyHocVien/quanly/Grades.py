from icecream import ic
from tabulate import tabulate
from openpyxl import Workbook
class QuanLyDiemThi():
    def __init__(self, db):
        self.cur = db.cur
        self.conn = db.conn

    def them(self):
        
        student_id = (input("Nhập mã học viên: "))
        course_id = (input("Nhập mã môn học: "))
        try: 
            list_1 = []
            list_2 = []
            sql = "SELECT id FROM students"
            self.cur.execute(sql)
            results = self.cur.fetchall()
            for i in results:
                list_1.append(str(i[0]))
            
            sql1 = "SELECT id FROM courses"
            self.cur.execute(sql1)
            results1 = self.cur.fetchall()
            for i in results1:
                list_2.append(str(i[0]))
            
            if student_id in list_1  and course_id in list_2:
                midterm_score = float(input("Nhập điểm quá trình: "))
                final_score = float(input("Nhập điểm kết thúc: "))

                sql = ("""
                INSERT INTO grades(student_id, course_id,midterm_score, final_score)
                VALUES (%s, %s, %s, %s)
                """)
                vals = [(student_id,course_id,midterm_score,final_score)]
                self.cur.executemany(sql, vals)
                self.conn.commit()

                print("Nhập điểm thành công.")
            else : 
                print(f"Không tồn tại mã  học viên và môn học  : {student_id}, {course_id}")
        except:
            print("Du lieu dau vao khong hop le!")    

    def sua(self):
        
        student_id = input("Nhập mã học viên muốn sửa điểm: ")
        course_id = input("Nhập mã môn học muốn sửa điểm: ")
        try:
            sql = """SELECT * FROM grades 
                    JOIN students on students.id = grades.student_id
                    JOIN courses on courses.id = grades.course_id
                    WHERE students.id = %s and courses.id = %s"""
            value = [student_id,course_id]
            self.cur.execute(sql, value)

            result = self.cur.fetchone()

            if result is None:
                print("Không tìm thấy học viên.")
                return
            midterm_score = input("Nhập điểm quá trình: ")
            final_score = input("Nhập điểm kết thúc: ")

            # Thực hiện truy vấn UPDATE vào database
            sql = "UPDATE grades SET midterm_score = %s, final_score = %s WHERE student_id = %s AND course_id = %s"
            values = (midterm_score, final_score,student_id,course_id )
            self.cur.execute(sql, values)
            self.conn.commit()

            print("Sửa điểm thành công.")
        except:
            print("Du lieu dau vao khong hop le!")
    def xoa(self):
        id = input("Nhập mã học viên muốn xoá điểm thi: ")
        try:
            sql = """SELECT * FROM grades
                    JOIN students on students.id = grades.student_id
                    WHERE students.id =%s"""
            value = [id]
            self.cur.execute(sql, value)
            result = self.cur.fetchone()
            if result is None:
                print("Không tìm thấy học viên.")
                return


            sql = """DELETE FROM grades
                    JOIN students on students.id = grades.student_id
                    WHERE students.id =%s"""
            value = [id]
            self.cur.execute(sql, value)
            self.conn.commit()

            print("Xoá điểm thành công.")
        except:
            print("Du lieu dau vao khong hop le!")

    def timKiem(self):
        keyword = input("Nhập mã học viên hoặc họ tên học viên muốn tra cứu điểm: ")
        try:
            if self.check_id(keyword):
                id = keyword
                # Thực hiện truy vấn SELECT vào database
                sql = """SELECT * FROM grades
                        JOIN students on students.id = grades.student_id
                        WHERE students.id =%s"""
                value = [id]
                
            else:
                name = keyword
                # Thực hiện truy vấn SELECT vào database
                sql = """SELECT * FROM grades
                        JOIN students on students.id = grades.student_id
                        WHERE students.name =%s"""
                value = [name]

            self.cur.execute(sql, value)
            result = self.cur.fetchall()

            if result is None:
                print("Không tìm thấy học viên.")
                return
                
            self.tabulate_print(result)
        except:
            print("Du lieu dau vao khong hop le!")
        
    def thongKe(self):
        print("Thống kê điểm: ")
        sql = """SELECT students.id,grades.midterm_score,grades.final_score ,(grades.midterm_score + grades.final_score*2)/3 as TongDiem,
                CASE 
                    when 90<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3<= 100 then "A"
                    when 70< (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 90 then "B"
                    when 50<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 70 then "C"
                    when 0<= (grades.midterm_score + grades.final_score*2)/3 AND  (grades.midterm_score + grades.final_score*2)/3< 50 then "D"
                    else "Khong phu hop"
                    END as LOAI
                FROM grades
                JOIN students on students.id = grades.student_id;
                """
        self.cur.execute(sql)
        results = self.cur.fetchall()
        if results is None:
            print("Không tìm thấy học viên.")
            return
        else:
            self.tabulate_print_tong_diem(results)
        
        sql1 = """SELECT COUNT(CASE when 90<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3<= 100 then "A" ELSE NULL END) AS LOAI_A,
                        COUNT(CASE when 70<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 90 then "B" ELSE NULL END) AS LOAI_B,
                        COUNT(CASE when 50<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 70 then "C" ELSE NULL END) AS LOAI_C,
                        COUNT(CASE when 0<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 50 then "D" ELSE NULL END) AS LOAI_D

                FROM grades
                JOIN students on students.id = grades.student_id
                JOIN courses on courses.id = grades.course_id
        """
        self.cur.execute(sql1)
        results1 = self.cur.fetchall()
        self.tabulate_print_so_luong(results1)


    def export_all_grades_to_excel(self):
        try:
            sql = """SELECT students.*,courses.*,grades.midterm_score,grades.final_score ,(grades.midterm_score + grades.final_score*2)/3 as TongDiem,
                    CASE 
                        when 90<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3<= 100 then "A"
                        when 70< (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 90 then "B"
                        when 50<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 70 then "C"
                        when 0<= (grades.midterm_score + grades.final_score*2)/3 AND  (grades.midterm_score + grades.final_score*2)/3< 50 then "D"
                        else "Khong phu hop"
                        END
                    FROM grades
                    JOIN students on students.id = grades.student_id
                    JOIN courses on courses.id = grades.course_id
                    
                    """
            self.cur.execute(sql)
            results = self.cur.fetchall()
            table_name = [i[0] for i in self.cur.description]
            print(table_name)
            print(results)

            wb =  Workbook()
            ws = wb.active
            ws.title = "mysql_data"
            ws.append(table_name)
            for row in results:
                ws.append(row)
            wb.save("All Greades.xlsx")
            self.conn.commit()
        except:
            print("Loi")

    def export_grade_A_to_excel(self):
        try:
            sql = """SELECT students.*,courses.*,grades.midterm_score,grades.final_score ,(grades.midterm_score + grades.final_score*2)/3 as TongDiem 
                    FROM grades
                    JOIN students on students.id = grades.student_id
                    JOIN courses on courses.id = grades.course_id
                    WHERE 90<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3<= 100 
                    """
            self.cur.execute(sql)
            results = self.cur.fetchall()
            table_name = [i[0] for i in self.cur.description]
            print(table_name)
            print(results)

            wb =  Workbook()
            ws = wb.active
            ws.title = "mysql_data"
            ws.append(table_name)
            for row in results:
                ws.append(row)
            wb.save("Grades A.xlsx")
            self.conn.commit()
        except:
            print("Loi")


    def export_grade_B_to_excel(self):
        try:
            sql = """SELECT students.*,courses.*,grades.midterm_score,grades.final_score ,(grades.midterm_score + grades.final_score*2)/3 as TongDiem 
                    FROM grades
                    JOIN students on students.id = grades.student_id
                    JOIN courses on courses.id = grades.course_id
                    WHERE 70< (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 90 
                    """
            self.cur.execute(sql)
            results = self.cur.fetchall()
            table_name = [i[0] for i in self.cur.description]
            print(table_name)
            print(results)

            wb =  Workbook()
            ws = wb.active
            ws.title = "mysql_data"
            ws.append(table_name)
            for row in results:
                ws.append(row)
            wb.save("Grades B.xlsx")
            self.conn.commit()
        except:
            print("Loi")


    def export_grade_C_to_excel(self):
        try:
            sql = """SELECT students.*,courses.*,grades.midterm_score,grades.final_score ,(grades.midterm_score + grades.final_score*2)/3 as TongDiem 
                    FROM grades
                    JOIN students on students.id = grades.student_id
                    JOIN courses on courses.id = grades.course_id
                    WHERE 50<= (grades.midterm_score + grades.final_score*2)/3 AND (grades.midterm_score + grades.final_score*2)/3< 70 
                    """
            self.cur.execute(sql)
            results = self.cur.fetchall()
            table_name = [i[0] for i in self.cur.description]
            print(table_name)
            print(results)

            wb =  Workbook()
            ws = wb.active
            ws.title = "mysql_data"
            ws.append(table_name)
            for row in results:
                ws.append(row)
            wb.save("Grades C.xlsx")
            self.conn.commit()
        except:
            print("Loi")


    def export_grade_D_to_excel(self):
        try:
            sql = """SELECT students.*,courses.*,grades.midterm_score,grades.final_score ,(grades.midterm_score + grades.final_score*2)/3 as TongDiem 
                    FROM grades
                    JOIN students on students.id = grades.student_id
                    JOIN courses on courses.id = grades.course_id
                    WHERE 0<= (grades.midterm_score + grades.final_score*2)/3 AND  (grades.midterm_score + grades.final_score*2)/3< 50 
                    """
            self.cur.execute(sql)
            results = self.cur.fetchall()
            table_name = [i[0] for i in self.cur.description]
            print(table_name)
            print(results)

            wb =  Workbook()
            ws = wb.active
            ws.title = "mysql_data"
            ws.append(table_name)
            for row in results:
                ws.append(row)
            wb.save("Grades D.xlsx")
            self.conn.commit()
        except:
            print("Loi")

    def check_id(self, id):
        try:
            id = int(id)
            return True
        except:
            print("")

        return False
    
    def hienThi(self):
        print("Danh sách điểm thi:")
        sql = "SELECT * FROM grades"
        self.cur.execute(sql)
        results = self.cur.fetchall()
        if results is None:
            print("Không tìm thấy điểm thi.")
            return
        
        self.tabulate_print(results)

    def tabulate_print(self, data):
        headers = ["Mã học viên", "Mã môn học", "Điểm quá trình", "Điểm kết thúc"]
        print(tabulate(data, headers, tablefmt="grid"))
    
    def tabulate_print_tong_diem(self, data):
        headers = ["Mã học viên", "Điểm quá trình", "Điểm kết thúc","Điểm tổng kết","Loại"]
        print(tabulate(data, headers, tablefmt="grid"))
    def tabulate_print_so_luong(self, data):
        headers = ["Loai A", "Loai B", "Loai C","Loai D"]
        print(tabulate(data, headers, tablefmt="grid"))